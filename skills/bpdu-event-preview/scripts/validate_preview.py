#!/usr/bin/env python3
"""
BP Debate Union — Event Preview Excel Validator

Validates an event preview Excel file against school standards.
Identifies problems and proposes fixes. Fixes are NOT applied
automatically — user is asked to approve each fix.

Usage:
    python validate_preview.py "path/to/file.xlsx" --week 10

Exit codes:
    0 = all pass
    1 = problems found (fix proposals printed)
"""

import argparse
import os
import re
import sys

import pandas as pd

VALID_ACTIVITY_TYPES = {"文化沙龙", "日常训练", "常规活动"}
EXPECTED_CLUB_NAME = "BP Debate Union"
TITLE_PATTERN = re.compile(
    r"温州商学院\d{4}-\d{4}学年第[一二三四五六七八九十百\d]+学期第[一二三四五六七八九十百\d]+周社团活动预告"
)
TIME_PATTERN = re.compile(
    r"\d{4}年\d{1,2}月\d{1,2}日 \d{1,2}:\d{2}-\d{1,2}:\d{2}"
)
LOCATION_PATTERN = re.compile(r"博闻楼[B研]-[A-Za-z0-9]+")


def validate(filepath, expected_week=None):
    """Returns (passed: bool, problems: list[dict])."""
    problems = []

    if not os.path.exists(filepath):
        return False, [{"type": "file", "message": f"File not found: {filepath}"}]

    expected_cols = {"社团名称", "活动内容", "活动地点", "开展时间"}

    try:
        # First, read raw (no header) to detect file structure
        df_raw = pd.read_excel(filepath, header=None)

        # Use openpyxl to check visual styles (fonts, sizes, borders)
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active

        # Try header=0: row 0 = title, row 1 = column names, row 2+ = data
        df = pd.read_excel(filepath, header=0)
        actual_cols = set(df.columns)
        title_row_idx = 0  # With header=0, title is at df index 0 (raw row 0)

        header_row = 0  # tracks which pandas header row was used
        if actual_cols != expected_cols:
            # Try header=1: row 0 = title (merged), row 1 = column names, row 2+ = data
            # Note: with header=1, pandas skips row 0, so df contains only data rows.
            df = pd.read_excel(filepath, header=1)
            actual_cols = set(df.columns)
            header_row = 1

        if actual_cols != expected_cols:
            return False, [{"type": "structure", "message": f"Column headers do not match expected: {expected_cols}. Found: {actual_cols}. File may have an unexpected structure."}]

        # Title is always at raw row 0 regardless of header setting
        title_text = str(df_raw.iloc[0, 0]) if len(df_raw) > 0 else ""

        # Check if row 1 is merged (A1:D1)
        is_merged = False
        for merged_range in ws.merged_cells.ranges:
            if "A1" in merged_range and "D1" in merged_range:
                is_merged = True
                break

        if not is_merged:
            problems.append({
                "type": "style",
                "field": "Row 1",
                "severity": "warning",
                "message": "Title row (A1:D1) should be merged"
            })

        # ---- 0. Visual Style Validation (openpyxl) ----
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4), start=1):
            for cell in row:
                # Skip secondary cells in the merged title range to avoid false positives
                if r_idx == 1 and cell.column > 1 and is_merged:
                    continue

                # Border check
                if not cell.border or not (cell.border.left.style and cell.border.right.style and cell.border.top.style and cell.border.bottom.style):
                    problems.append({
                        "type": "style",
                        "field": f"Cell {cell.coordinate}",
                        "severity": "warning",
                        "message": f"Cell {cell.coordinate} is missing borders"
                    })

                # Font check
                if not cell.font or cell.font.name != '等线':
                    problems.append({
                        "type": "style",
                        "field": f"Cell {cell.coordinate}",
                        "severity": "warning",
                        "message": f"Cell {cell.coordinate} font is not '等线' (found: {cell.font.name if cell.font else 'None'})"
                    })

                # Size and Bold check
                if r_idx == 1: # Title
                    if cell.font and (cell.font.size != 22 or cell.font.bold):
                        problems.append({
                            "type": "style",
                            "field": f"Cell {cell.coordinate} (Title)",
                            "severity": "warning",
                            "message": f"Title font should be size 22 and NOT bold"
                        })
                elif r_idx == 2: # Header
                    if cell.font and (cell.font.size != 11 or cell.font.bold):
                        problems.append({
                            "type": "style",
                            "field": f"Cell {cell.coordinate} (Header)",
                            "severity": "warning",
                            "message": f"Header font should be size 11 and NOT bold"
                        })
                else: # Body
                    if cell.font and (cell.font.size != 11 or cell.font.bold):
                        problems.append({
                            "type": "style",
                            "field": f"Cell {cell.coordinate} (Body)",
                            "severity": "warning",
                            "message": f"Body font should be size 11 and NOT bold"
                        })

                # Alignment check
                if not cell.alignment or cell.alignment.horizontal != 'center' or cell.alignment.vertical != 'center':
                    problems.append({
                        "type": "style",
                        "field": f"Cell {cell.coordinate}",
                        "severity": "warning",
                        "message": f"Cell {cell.coordinate} is not center-aligned"
                    })

    except Exception as e:
        return False, [{"type": "file", "message": f"Cannot read Excel file: {e}"}]

    # Ensure we have at least 4 columns
    if df.shape[1] < 4:
        problems.append({
            "type": "structure",
            "field": "columns",
            "severity": "error",
            "message": f"Expected at least 4 columns, found {df.shape[1]}"
        })
        return False, problems

    # ---- 1. Title row ----
    if not TITLE_PATTERN.search(title_text):
        problems.append({
            "type": "title",
            "field": "社团名称 (row 1)",
            "severity": "error",
            "message": f"Title row format incorrect: '{title_text}'"
        })

    # ---- 2. Data rows ----
    # With header=0: df.iloc[0] is the title row (carried as a data row by pandas), skip it.
    # With header=1: pandas already excluded the title row; all of df is data.
    data_rows = df.iloc[1:].reset_index(drop=True) if header_row == 0 else df.reset_index(drop=True)

    if data_rows.empty:
        problems.append({
            "type": "data",
            "field": "rows",
            "severity": "error",
            "message": "No activity data rows found"
        })

    # ---- 4. Validate each data row ----
    week_mismatches = []
    for i, row in data_rows.iterrows():
        row_num = i + 2  # Excel row number (1-indexed, title is row 1)

        # 社团名称
        if str(row["社团名称"]).strip() != EXPECTED_CLUB_NAME:
            problems.append({
                "type": "data",
                "field": f"社团名称 (row {row_num})",
                "severity": "error",
                "message": f"Expected '{EXPECTED_CLUB_NAME}', found '{row['社团名称']}'"
            })

        # 活动内容
        content = str(row["活动内容"]).strip()
        if content not in VALID_ACTIVITY_TYPES:
            problems.append({
                "type": "data",
                "field": f"活动内容 (row {row_num})",
                "severity": "warning",
                "message": f"Activity type '{content}' not in standard list {VALID_ACTIVITY_TYPES}"
            })

        # 活动地点
        location = str(row["活动地点"]).strip()
        # Only validate building+room format when location looks like one.
        # If it contains building-related keywords, warn if spaces are present or format is wrong.
        # Arbitrary locations (e.g., "线上", "博闻楼门口集合") are always accepted.
        if any(kw in location for kw in ("博闻楼", "研", "楼", "室", "馆")):
            if " " in location or not LOCATION_PATTERN.search(location):
                problems.append({
                    "type": "data",
                    "field": f"活动地点 (row {row_num})",
                    "severity": "warning",
                    "message": f"Location '{location}' appears to be building+room but has a space or wrong format"
                })

        # 开展时间 format
        time_str = str(row["开展时间"]).strip()
        if not TIME_PATTERN.match(time_str):
            problems.append({
                "type": "data",
                "field": f"开展时间 (row {row_num})",
                "severity": "error",
                "message": f"Time format incorrect: '{time_str}'"
            })

        # Week number match (if expected_week given)
        if expected_week is not None:
            time_match = TIME_PATTERN.search(time_str)
            if time_match:
                date_str = time_match.group()
                # Extract week from title
                week_match = re.search(r"第([一二三四五六七八九十百\d]+)周", title_text)
                if week_match:
                    cn_nums = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5,
                               "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}
                    week_cn = week_match.group(1)
                    try:
                        actual_week = int(week_cn) if week_cn.isdigit() else cn_nums.get(week_cn[0], None)
                        if actual_week != expected_week:
                            week_mismatches.append((row_num, actual_week, expected_week))
                    except (ValueError, KeyError):
                        pass

    if week_mismatches:
        for row_num, actual, expected in week_mismatches:
            problems.append({
                "type": "week_mismatch",
                "field": f"title / row {row_num}",
                "severity": "error",
                "message": f"Title declares week {actual}, but --week={expected} was expected"
            })

    passed = len(problems) == 0
    return passed, problems


def print_report(passed, problems):
    print("\n=== Validation Results ===")
    if passed:
        print("All checks PASSED.")
        return

    errors = [p for p in problems if p.get("severity") == "error"]
    warnings = [p for p in problems if p.get("severity") == "warning"]

    for p in errors + warnings:
        label = "[ERROR]" if p["severity"] == "error" else "[WARNING]"
        print(f"\n{label} {p['message']}")
        print(f"  Location: {p.get('field', 'N/A')}")

    print(f"\n--- Summary ---")
    print(f"Errors: {len(errors)}  |  Warnings: {len(warnings)}")


def main():
    parser = argparse.ArgumentParser(description="Validate BP DU Event Preview Excel.")
    parser.add_argument("xlsx_path", help="Path to the event preview .xlsx file")
    parser.add_argument("--week", type=int, default=None,
                        help="Expected week number (e.g. 10 for 第十周)")
    args = parser.parse_args()

    passed, problems = validate(args.xlsx_path, expected_week=args.week)
    print_report(passed, problems)

    sys.exit(0 if passed else 1)


if __name__ == "__main__":
    main()
