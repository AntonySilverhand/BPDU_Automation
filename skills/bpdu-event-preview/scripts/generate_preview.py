#!/usr/bin/env python3
"""
BP Debate Union — Event Preview Excel Generator

Generates a weekly activity preview Excel report.

Usage:
    python generate_preview.py --week 10
    # Or edit INPUT DATA below

All generation scripts write to ../output/ (relative to this script's location).
"""

import os
import argparse
from datetime import datetime

import pandas as pd

# ===================== INPUT DATA =====================
WEEK_NUMBER = 7
ACTIVITIES = [
    {"日期": "2026年4月14日",  "时间": "18:20-20:20", "内容": "常规活动", "地点": "博闻楼B-602"},
    {"日期": "2026年4月15日",  "时间": "18:20-20:20", "内容": "常规活动", "地点": "博闻楼B-604"},
]
# =====================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(SCRIPT_DIR)
OUTPUT_DIR = os.path.join(SKILL_DIR, "output")


def parse_args():
    parser = argparse.ArgumentParser(description="Generate BP DU Event Preview Excel.")
    parser.add_argument("--week", type=int, default=WEEK_NUMBER,
                        help="Week number (e.g. 7 for 第七周)")
    parser.add_argument("--activity", action="append", nargs=4,
                        metavar=("DATE", "TIME", "CONTENT", "LOCATION"),
                        help="Add an activity. Can be repeated. "
                             "Format: --activity '2026年4月14日' '18:20-20:20' '常规活动' '博闻楼B-602'")
    return parser.parse_args()


def make_unique_path(filepath):
    """Append timestamp if file already exists to avoid overwrite."""
    if not os.path.exists(filepath):
        return filepath
    base, ext = os.path.splitext(filepath)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"{base}_{timestamp}{ext}"


def generate(args):
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    activities_to_process = []
    if args.activity:
        for act in args.activity:
            activities_to_process.append({
                "日期": act[0],
                "时间": act[1],
                "内容": act[2],
                "地点": act[3]
            })
    else:
        activities_to_process = ACTIVITIES

    data = []
    for a in activities_to_process:
        data.append({
            "社团名称": "BP Debate Union",
            "活动内容": a["内容"],
            "活动地点": a["地点"],
            "开展时间": f"{a['日期']} {a['时间']}"
        })

    df = pd.DataFrame(data)

    # Title row text
    title = f"温州商学院2025-2026学年第二学期第{args.week}周社团活动预告"

    filename = f"BP_Debate_Union_第{args.week}周活动预告汇总.xlsx"
    filepath = make_unique_path(os.path.join(OUTPUT_DIR, filename))

    # Write df starting from row 2 (startrow=1 is 0-indexed, so Row 2)
    # This puts headers on Row 2 and data on Row 3+
    df.to_excel(filepath, index=False, engine='openpyxl', startrow=1)

    # Post-processing with openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Side
    wb = load_workbook(filepath)
    ws = wb.active

    # 1. Add title to Row 1 and merge
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # 2. Styles: Center all cells, Auto-fit columns, and Add Borders
    from openpyxl.styles import Font
    center_align = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Calculate widths while applying styles
    column_widths = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4), start=1):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border
            # Bold the header row (Row 2) and Title (Row 1)
            if row_idx <= 2:
                cell.font = bold_font

            if cell.value:
                # Estimate width: Chinese characters count as 2, others as 1
                val_str = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val_str)
                col_letter = cell.column_letter
                if length > column_widths.get(col_letter, 0):
                    column_widths[col_letter] = length

    # Apply widths
    for col_letter, width in column_widths.items():
        # Add a little padding
        ws.column_dimensions[col_letter].width = width + 2

    wb.save(filepath)

    print(f"Generated: {filepath}")
    print("\nFile contents (data only):")
    print(df.to_string(index=False))
    return filepath


if __name__ == "__main__":
    args = parse_args()
    generate(args)
